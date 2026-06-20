"""Text, CSV, and spreadsheet loader.

Purpose
-------
Reads simple local formats (plain text, Markdown, JSON, CSV, and Excel) into
normalized records.

What it does
------------
Loads text files with encoding fallbacks, parses CSV and Excel into per-row
records, and converts each row into readable text suitable for chunking.

Flow
----
The loader picks a reader based on file type, extracts rows or text, and emits one
record per row (for tabular data) or per file (for text), each ready for the
chunker.
"""

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.app.ingestion.loaders.base import LoadedFile, LoadedRecord
from backend.app.ingestion.metadata import short_hash


class LocalFileLoader:
    """
    Compact loader for simple local file formats.

    This class extracts records only. It does not:
        - create chunks directly
        - write SQLite
        - create embeddings
        - index vectors
        - call the LLM
    """

    supported_extensions = {
        ".txt",
        ".md",
        ".markdown",
        ".csv",
        ".xlsx",
    }

    text_extensions = {
        ".txt",
        ".md",
        ".markdown",
    }

    csv_extensions = {
        ".csv",
    }

    xlsx_extensions = {
        ".xlsx",
    }

    def __init__(self, encodings: list[str] | None = None):
        self.encodings = encodings or ["utf-8", "utf-8-sig", "cp1252"]

    def supports(self, path: str | Path) -> bool:
        """
        Return True if this loader supports the file extension.
        """

        return Path(path).suffix.lower() in self.supported_extensions

    def load(self, path: str | Path) -> LoadedFile:
        """
        Load a supported local file and return normalized records.

        Args:
            path:
                Path to the source file.

        Returns:
            LoadedFile:
                Normalized file-level and record-level output.
        """

        file_path = Path(path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = file_path.suffix.lower()

        if suffix in self.text_extensions:
            return self._load_text_file(file_path)

        if suffix in self.csv_extensions:
            return self._load_csv_file(file_path)

        if suffix in self.xlsx_extensions:
            return self._load_xlsx_file(file_path)

        raise ValueError(f"Unsupported file type: {suffix}")

    def _load_text_file(self, path: Path) -> LoadedFile:
        """
        Load TXT/MD/Markdown as a single text document record.
        """

        text, encoding = self._read_text_with_fallback(path)

        return LoadedFile(
            source_path=path,
            records=[
                LoadedRecord(
                    text=text,
                    record_type="text_document",
                    title=path.name,
                    metadata={
                        "encoding": encoding,
                        "file_name": path.name,
                        "file_suffix": path.suffix.lower(),
                    },
                )
            ],
            document_metadata={
                "encoding": encoding,
                "file_name": path.name,
                "file_suffix": path.suffix.lower(),
                "loader": "LocalFileLoader",
            },
        )

    def _load_csv_file(self, path: Path) -> LoadedFile:
        """
        Load CSV as one record per row.

        Each row is converted into readable field-labeled text so that both
        keyword retrieval and future semantic retrieval work well.
        """

        text, encoding = self._read_text_with_fallback(path)
        rows, columns = self._parse_csv_text(text, path)

        loaded_records = self._rows_to_loaded_records(
            rows=rows,
            columns=columns,
            path=path,
            file_suffix=".csv",
            source_kind="csv",
            encoding=encoding,
        )

        return LoadedFile(
            source_path=path,
            records=loaded_records,
            document_metadata={
                "encoding": encoding,
                "file_name": path.name,
                "file_suffix": path.suffix.lower(),
                "loader": "LocalFileLoader",
                "columns": columns,
                "row_count": len(rows),
            },
        )

    def _load_xlsx_file(self, path: Path) -> LoadedFile:
        """
        Load XLSX as one record per non-empty row per worksheet.

        The first non-empty row in each sheet is treated as the header row.
        Each subsequent non-empty row becomes one LoadedRecord.
        """

        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )

        all_records: list[LoadedRecord] = []
        sheet_summaries: list[dict[str, Any]] = []

        for worksheet in workbook.worksheets:
            rows = self._worksheet_to_rows(worksheet)

            if not rows:
                sheet_summaries.append(
                    {
                        "sheet_name": worksheet.title,
                        "columns": [],
                        "row_count": 0,
                    }
                )
                continue

            columns = list(rows[0].keys())

            sheet_records = self._rows_to_loaded_records(
                rows=rows,
                columns=columns,
                path=path,
                file_suffix=".xlsx",
                source_kind="xlsx",
                encoding=None,
                sheet_name=worksheet.title,
            )

            all_records.extend(sheet_records)

            sheet_summaries.append(
                {
                    "sheet_name": worksheet.title,
                    "columns": columns,
                    "row_count": len(rows),
                }
            )

        workbook.close()

        return LoadedFile(
            source_path=path,
            records=all_records,
            document_metadata={
                "file_name": path.name,
                "file_suffix": path.suffix.lower(),
                "loader": "LocalFileLoader",
                "sheets": sheet_summaries,
                "sheet_count": len(sheet_summaries),
                "row_count": sum(sheet["row_count"] for sheet in sheet_summaries),
            },
        )

    def _read_text_with_fallback(self, path: Path) -> tuple[str, str]:
        """
        Read text using the first successful encoding.
        """

        last_error: UnicodeDecodeError | None = None

        for encoding in self.encodings:
            try:
                return path.read_text(encoding=encoding), encoding
            except UnicodeDecodeError as exc:
                last_error = exc

        if last_error:
            raise last_error

        raise UnicodeDecodeError(
            "unknown",
            b"",
            0,
            1,
            f"Unable to decode file: {path}",
        )

    @staticmethod
    def _parse_csv_text(text: str, path: Path) -> tuple[list[dict[str, str]], list[str]]:
        """
        Parse CSV text using csv.DictReader.
        """

        lines = text.splitlines()
        reader = csv.DictReader(lines)

        if not reader.fieldnames:
            raise ValueError(f"CSV file has no header row: {path}")

        columns = [column.strip() for column in reader.fieldnames]
        rows: list[dict[str, str]] = []

        for raw_row in reader:
            cleaned_row = {
                str(key).strip(): str(value).strip() if value is not None else ""
                for key, value in raw_row.items()
                if key is not None
            }

            if any(value for value in cleaned_row.values()):
                rows.append(cleaned_row)

        return rows, columns

    @staticmethod
    def _worksheet_to_rows(worksheet: Any) -> list[dict[str, str]]:
        """
        Convert one worksheet into a list of row dictionaries.

        The first non-empty row is treated as the header row.
        """

        raw_rows = list(worksheet.iter_rows(values_only=True))

        non_empty_rows = [
            row
            for row in raw_rows
            if row is not None and any(cell is not None and str(cell).strip() for cell in row)
        ]

        if not non_empty_rows:
            return []

        header_row = non_empty_rows[0]
        headers = [
            str(cell).strip() if cell is not None and str(cell).strip() else f"column_{index + 1}"
            for index, cell in enumerate(header_row)
        ]

        rows: list[dict[str, str]] = []

        for raw_row in non_empty_rows[1:]:
            row: dict[str, str] = {}

            for index, header in enumerate(headers):
                value = raw_row[index] if index < len(raw_row) else ""
                row[header] = LocalFileLoader._cell_to_text(value)

            if any(value for value in row.values()):
                rows.append(row)

        return rows

    @staticmethod
    def _cell_to_text(value: Any) -> str:
        """
        Convert an XLSX cell value into clean text.
        """

        if value is None:
            return ""

        return str(value).strip()

    def _rows_to_loaded_records(
        self,
        rows: list[dict[str, str]],
        columns: list[str],
        path: Path,
        file_suffix: str,
        source_kind: str,
        encoding: str | None = None,
        sheet_name: str | None = None,
    ) -> list[LoadedRecord]:
        """
        Convert structured rows into LoadedRecord objects.

        This shared method is used for CSV and XLSX.
        """

        loaded_records: list[LoadedRecord] = []

        for index, row in enumerate(rows):
            row_text = self._row_to_text(row)

            if not row_text:
                continue

            row_number = index + 2
            row_hash = short_hash(json.dumps(row, sort_keys=True, ensure_ascii=False))
            title = self._choose_row_title(row, fallback=f"{path.name} row {row_number}")

            metadata = {
                "file_name": path.name,
                "file_suffix": file_suffix,
                "row_number": row_number,
                "row_hash": row_hash,
                "columns": columns,
                "raw_row": row,
            }

            if encoding:
                metadata["encoding"] = encoding

            if sheet_name:
                metadata["sheet_name"] = sheet_name

            loaded_records.append(
                LoadedRecord(
                    text=row_text,
                    record_type=f"{source_kind}_row",
                    title=title,
                    metadata=metadata,
                )
            )

        return loaded_records

    @staticmethod
    def _row_to_text(row: dict[str, str]) -> str:
        """
        Convert a structured row into readable text.

        Example:
            asset_id: C-900
            error_code: E-112

        Becomes:
            Asset ID: C-900
            Error Code: E-112
        """

        lines: list[str] = []

        for key, value in row.items():
            if value == "":
                continue

            label = LocalFileLoader._normalize_column_label(key)
            lines.append(f"{label}: {value}")

        return "\n".join(lines).strip()

    @staticmethod
    def _normalize_column_label(column_name: str) -> str:
        """
        Convert a structured column name into a readable label.
        """

        cleaned = column_name.replace("_", " ").replace("-", " ").strip()
        return cleaned.title() if cleaned else "Unknown Field"

    @staticmethod
    def _choose_row_title(row: dict[str, str], fallback: str) -> str:
        """
        Pick a useful title for a structured row.
        """

        candidate_fields = [
            "title",
            "name",
            "asset_id",
            "equipment_name",
            "machine_id",
            "employee_id",
            "employee_name",
            "group_id",
            "group_name",
            "error_code",
        ]

        lowered = {key.lower(): value for key, value in row.items()}

        for field in candidate_fields:
            value = lowered.get(field)
            if value:
                return value

        return fallback